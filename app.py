import streamlit as st
import pandas as pd
from io import BytesIO
import re
from openpyxl import load_workbook

# 页面配置
st.set_page_config(page_title="BioMuse 订单自动化助手", layout="wide")
# --- 隐藏所有 GitHub 痕迹的样式 ---
st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    /* 这一行会彻底隐藏右上角的 GitHub 图标 */
    .stAppDeployButton {display:none;} 
    </style>
    """, unsafe_allow_html=True)
st.title("🧬 BioMuse 订单自动化助手")
st.caption("科研销售效率工具：自动识别序列并填充百力格订购表")

# --- 核心逻辑函数 ---

def get_complement(seq, is_rna=True):
    """生成互补序列并反转 (5'->3')"""
    # 统一转大写
    seq = seq.upper().replace(' ', '')
    # 定义配对字典
    if is_rna:
        pairs = {'A': 'U', 'U': 'A', 'G': 'C', 'C': 'G', 'T': 'A'}
    else:
        pairs = {'A': 'T', 'T': 'A', 'G': 'C', 'C': 'G'}
    
    complement = "".join([pairs.get(base, 'N') for base in seq])
    return complement[::-1]  # 反转得到 5' to 3'

def process_rna(sense_seq):
    """RNA 专用规则：生成 Antisense 并加 dTdT"""
    # 1. 生成互补反向链
    antisense = get_complement(sense_seq, is_rna=True)
    # 2. 根据百力格规范加 dTdT (19+2 模式)
    return sense_seq + "dTdT", antisense + "dTdT"

# --- 界面交互 ---

st.header("1. 客户基本信息")
# 创建两列，让输入框并排，节省手机空间
col1, col2 = st.columns(2)
with col1:
    c_name = st.text_input("客户姓名", "待填写")
with col2:
    c_unit = st.text_input("客户单位", "待填写")
c_group = st.text_input("课题组", "待填写")
order_type = st.radio("选择订单模式", ["DNA引物", "siRNA/RNA"], horizontal=True)

st.header("2. 粘贴客户需求")
raw_text = st.text_area("直接粘贴客户发来的文字（包含名称、序列、OD等）：", height=200, 
                       placeholder="例如：\nF: ATGC...\nR: TCCG...\n2OD PAGE纯化")

if st.button("🚀 开始解析并生成表格"):
    if not raw_text:
        st.error("请输入需求内容！")
    else:
        # 解析数据
        data_rows = []
        lines = raw_text.strip().split('\n')
        
        if order_type == "DNA引物":
            # 简单正则提取 DNA
            matches = re.findall(r'([FR\d\w-]+)[:\s]+([ATCGatcg]+)', raw_text)
            for name, seq in matches:
                data_rows.append({
                    '引物名称': name, '引物序列': seq.upper(), 
                    '纯化方式': 'PAGE', '分装OD': 2
                })
        else:
            # RNA 模式逻辑
            matches = re.findall(r'(Sense|正义链)?[:\s]*([AUCGaucg]+)', raw_text)
            for _, seq in matches:
                s_final, a_final = process_rna(seq.upper())
                data_rows.append({
                    '引物名称': 'siRNA-S', '引物序列': s_final,
                    '反向名称': 'siRNA-A', '反向序列': a_final,
                    '纯化方式': 'HPLC', '分装OD': 2
                })

        if not data_rows:
            st.warning("未能识别出序列，请检查输入格式。")
        else:
            st.success(f"成功识别出 {len(data_rows)} 条序列！")
            st.table(pd.DataFrame(data_rows))

            # --- 填充 Excel 模板逻辑 ---
            template_path = "template_dna.xlsx" if order_type == "DNA引物" else "template_rna.xlsx"
            
            try:
                wb = load_workbook(template_path)
                ws = wb.active # 假设数据在第一个Sheet

                # 填充表头（根据百力格模板位置调整，这里假设在B列）
                ws['B5'] = client_name
                ws['B6'] = client_unit
                ws['B7'] = client_group

                # 填充数据行（从第20行开始填充）
                start_row = 20
                for i, row_data in enumerate(data_rows):
                    curr_row = start_row + i
                    if order_type == "DNA引物":
                        ws.cell(row=curr_row, column=2, value=row_data['引物名称'])
                        ws.cell(row=curr_row, column=3, value=row_data['引物序列'])
                        ws.cell(row=curr_row, column=4, value=row_data['纯化方式'])
                        ws.cell(row=curr_row, column=8, value=row_data['分装OD'])
                    else:
                        ws.cell(row=curr_row, column=1, value=row_data['引物名称'])
                        ws.cell(row=curr_row, column=2, value=row_data['引物序列'])
                        ws.cell(row=curr_row, column=3, value=row_data['反向序列'])
                        ws.cell(row=curr_row, column=4, value=row_data['纯化方式'])

                # 导出
                output = BytesIO()
                wb.save(output)
                st.download_button(
                    label="💾 点击下载生成的订购表",
                    data=output.getvalue(),
                    file_name=f"百力格订单_{client_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except FileNotFoundError:
                st.error(f"未找到模板文件 {template_path}，请确保它与 app.py 在同一目录下。")
